from __future__ import annotations

from collections import Counter
from io import BytesIO
from datetime import date, datetime
import re

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


APP_TITLE = "KDHW Children Verification"
SOURCE_SHEET = "children"
PW_SHEET = "PW"
CODE_HEADER = "children_code"
PW_CODE_HEADER = "mother_code"
VERIFICATION_SHEET = "children_verification"
LONG_FORM_SHEET = "children_long_form"
PW_VERIFICATION_SHEET = "pw_verification"
PW_LONG_FORM_SHEET = "pw_long"
VTHC_DISAGGREGATE_SHEET = "VTHC_dose_disaggregate"
INDICATOR_SHEET = "indicators"
CUMULATIVE_SHEET = "cummulative_sheet"
CUMMU_INDICATOR_SHEET = "Cummu_indicator"
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
RED_FONT = Font(color="9C0006")
SOURCE_FIELD_RULES = [
    ("BCG", "BCG_other", "BCG_source"),
    ("OPV_first_time_dose", "OPV_first_time_dose_other", "OPV1_source"),
    ("OPV_second_time_dose", "OPV_second_time_dose_other", "OPV2_source"),
    ("OPV_third_time_dose", "OPV_third_time_dose_other", "OPV3_source"),
    ("Penta_first_time_dose", "Penta_first_time_dose_other", "Penta1_source"),
    ("Penta_second_time_dose", "Penta_second_time_dose_other", "Penta2_source"),
    ("Penta_third_time_dose", "Penta_third_time_dose_other", "Penta3_source"),
    ("Penta_fourth_time_dose", "Penta_fourth_time_dose_other", "Penta4_source"),
    ("MMR_first_time_dose", "MMR_first_time_dose_other", "MMR1_source"),
    ("MMR_second_time_dose", "MMR_second_time_dose_other", "MMR2_source"),
    ("JE_first_time_dose", "JE_first_time_dose_other", "JE1_source"),
    ("JE_second_time_dose", "JE_second_time_dose_other", "JE2_source"),
    ("IPV", "IPV_other", "IPV_source"),
    ("Rota_first_time_dose", "Rota_first_time_dose_other", "Rota1_source"),
    ("Rota_second_time_dose", "Rota_second_time_dose_other", "Rota2_source"),
]
EXCLUDED_OUTPUT_HEADERS = {
    "HPV_first_time_dose",
    "HPV_first_time_dose_other",
    "HPV_first_time_dose_reporting_month",
    "HPV_second_time_dose",
    "HPV_second_time_dose_other",
    "HPV_second_time_dose_reporting_month",
    "HepB",
    "HepB_other",
    "HepB_reporting_month",
    "PCV_first_time_dose",
    "PCV_first_time_dose_other",
    "PCV_first_time_dose_reporting_month",
    "PCV_second_time_dose",
    "PCV_second_time_dose_other",
    "PCV_second_time_dose_reporting_month",
    "PCV_third_time_dose",
    "PCV_third_time_dose_other",
    "PCV_third_time_dose_reporting_month",
    "full_prevention",
    "full_prevention_reporting_month",
    "complete_dose",
    "complete_dose_reporting_month",
    "comments",
}
DATE_COLUMNS = [
    "registered_date",
    "BCG",
    "OPV_first_time_dose",
    "OPV_second_time_dose",
    "OPV_third_time_dose",
    "MMR_first_time_dose",
    "MMR_second_time_dose",
    "IPV",
    "PCV_first_time_dose",
    "PCV_second_time_dose",
    "PCV_third_time_dose",
    "Rota_first_time_dose",
    "Rota_second_time_dose",
]
DOSE_DATE_HEADERS = {
    "BCG": "BCG",
    "OPV1": "OPV_first_time_dose",
    "OPV2": "OPV_second_time_dose",
    "OPV3": "OPV_third_time_dose",
    "Penta1": "Penta_first_time_dose",
    "Penta2": "Penta_second_time_dose",
    "Penta3": "Penta_third_time_dose",
    "Penta4": "Penta_fourth_time_dose",
    "MMR1": "MMR_first_time_dose",
    "MMR2": "MMR_second_time_dose",
    "JE1": "JE_first_time_dose",
    "JE2": "JE_second_time_dose",
    "IPV": "IPV",
    "Rota1": "Rota_first_time_dose",
    "Rota2": "Rota_second_time_dose",
}
DOSE_SOURCE_HEADERS = {
    "BCG": "BCG_source",
    "OPV1": "OPV1_source",
    "OPV2": "OPV2_source",
    "OPV3": "OPV3_source",
    "Penta1": "Penta1_source",
    "Penta2": "Penta2_source",
    "Penta3": "Penta3_source",
    "Penta4": "Penta4_source",
    "MMR1": "MMR1_source",
    "MMR2": "MMR2_source",
    "JE1": "JE1_source",
    "JE2": "JE2_source",
    "IPV": "IPV_source",
    "Rota1": "Rota1_source",
    "Rota2": "Rota2_source",
}
DOSE_REPORTING_HEADERS = {
    "BCG": "BCG_reporting_month",
    "OPV1": "OPV_first_time_dose_reporting_month",
    "OPV2": "OPV_second_time_dose_reporting_month",
    "OPV3": "OPV_third_time_dose_reporting_month",
    "Penta1": "Penta_first_time_dose_reporting_month",
    "Penta2": "Penta_second_time_dose_reporting_month",
    "Penta3": "Penta_third_time_dose_reporting_month",
    "Penta4": "Penta_fourth_time_dose_reporting_month",
    "MMR1": "MMR_first_time_dose_reporting_month",
    "MMR2": "MMR_second_time_dose_reporting_month",
    "JE1": "JE_first_time_dose_reporting_month",
    "JE2": "JE_second_time_dose_reporting_month",
    "IPV": "IPV_reporting_month",
    "Rota1": "Rota_first_time_dose_reporting_month",
    "Rota2": "Rota_second_time_dose_reporting_month",
}
LONG_FORM_DOSES = [
    "BCG",
    "OPV1",
    "OPV2",
    "OPV3",
    "Penta1",
    "Penta2",
    "Penta3",
    "MMR1",
    "MMR2",
    "JE1",
    "JE2",
    "IPV",
    "Rota1",
    "Rota2",
]
U1_REQUIRED_DOSES = ["BCG", "OPV1", "OPV2", "OPV3", "Penta1", "Penta2", "Penta3", "MMR1"]
ONE_TO_FIVE_REQUIRED_DOSES = ["OPV1", "OPV2", "OPV3", "Penta1", "Penta2", "Penta3", "MMR1"]
COMPLETION_CHECK_DOSES = ["BCG", "OPV1", "OPV2", "OPV3", "Penta1", "Penta2", "Penta3", "MMR1", "MMR2"]
SEQUENCE_RULES = {
    "OPV": ["OPV1", "OPV2", "OPV3"],
    "Penta": ["Penta1", "Penta2", "Penta3"],
    "MMR": ["MMR1", "MMR2"],
}
PW_DROP_HEADERS = {
    "td_third_dose",
    "td_third_dose_other",
    "td_third_dose_reporting_month",
    "td_fourth_dose",
    "td_fourth_dose_other",
    "td_fourth_dose_reporting_month",
    "td_fifth_dose",
    "td_fifth_dose_other",
    "td_fifth_dose_reporting_month",
    "prevent_td_newborn",
    "comments",
}
PW_SOURCE_FIELD_RULES = [
    ("td_first_dose", "td_first_dose_other", "Td1_source"),
    ("td_second_dose", "td_second_dose_other", "Td2_source"),
]
PW_DOSE_DATE_HEADERS = {
    "Td1": "td_first_dose",
    "Td2": "td_second_dose",
}
PW_DOSE_REPORTING_HEADERS = {
    "Td1": "td_first_dose_reporting_month",
    "Td2": "td_second_dose_reporting_month",
}
PW_LONG_DOSES = ["Td1", "Td2"]
DISAGGREGATE_HEADERS = [
    "Year",
    "period",
    "Organization",
    "Project Name",
    "District (EHO)",
    "Township_EHO",
    "Twp_MIMU",
    "Clinic Name",
    "ALOD-U1",
    "ALOD-U5",
    "ALOD->5",
    "BCG_U1",
    "BCG_U5",
    "BCG_>5",
    "OPV1_U1",
    "OPV1_U5",
    "OPV1_>5",
    "OPV2_U1",
    "OPV2_U5",
    "OPV2_>5",
    "OPV3_U1",
    "OPV3_U5",
    "OPV3_>5",
    "Penta1_U1",
    "Penta1_U5",
    "Penta1_>5",
    "Penta2_U1",
    "Penta2_U5",
    "Penta2_>5",
    "Penta3_U1",
    "Penta3_U5",
    "Penta3_>5",
    "MMR1_U1",
    "MMR1_U5",
    "MMR1_>5",
    "MMR2_U1",
    "MMR2_U5",
    "MMR2_>5",
    "JE_U1",
    "JE_U5",
    "JE_>5",
    "IPV_U1",
    "IPV_U5",
    "IPV_>5",
    "CD_U1",
    "CD_U5",
    "CD_>5",
    "Td1",
    "Td2",
    "Td At least one dose",
]
DISAGGREGATE_COUNT_COLUMNS = DISAGGREGATE_HEADERS[8:]
YEARLY_CUMULATIVE_HEADERS = [
    "Organization",
    "period",
    "Project Name",
    "District (EHO)",
    "Township_EHO",
    "Twp_MIMU",
    "Clinic Name",
    "ALOD-U1",
    "ALOD-U5",
    "ALOD->5",
    "Td At least one dose",
]
INDICATOR_HEADERS = [
    "Year",
    "Organization",
    "Project Name",
    "indicator",
    "Q1 Target",
    "Q1 U1 Male",
    "Q1 U1 Female",
    "Q1 1-5 Male",
    "Q1 1-5 Female",
    "Q1 Total",
    "Q2 Target",
    "Q2 U1 Male",
    "Q2 U1 Female",
    "Q2 1-5 Male",
    "Q2 1-5 Female",
    "Q2 Total",
    "Q3 Target",
    "Q3 U1 Male",
    "Q3 U1 Female",
    "Q3 1-5 Male",
    "Q3 1-5 Female",
    "Q3 Total",
    "Q4 Target",
    "Q4 U1 Male",
    "Q4 U1 Female",
    "Q4 1-5 Male",
    "Q4 1-5 Female",
    "Q4 Total",
]
CUMMU_INDICATOR_HEADERS = [
    "Year",
    "Organization",
    "Project Name",
    "indicator",
    "Annual Target",
    "Annual U1 Male",
    "Annaul U1 Female",
    "Annual 1-5 Male",
    "Annual 1-5 Female",
]
INDICATOR_DEFINITIONS = [
    ("Penta3 under 1-yr-old", "Penta3", {"U1"}),
    ("MMR1 under 1-yr-old", "MMR1", {"U1"}),
    ("Penta1 under 5-yr-old", "Penta1", {"U1", "U5"}),
    ("Penta3 under 5-yr-old", "Penta3", {"U1", "U5"}),
    ("MMR1 under 5-yr-old", "MMR1", {"U1", "U5"}),
    ("MMR2 under 5-yr-old", "MMR2", {"U1", "U5"}),
    ("Full dose under 5-yr-old", "CD", {"U1", "U5"}),
    ("At least one dose under 5-yr-old", "ALOD", {"U1", "U5"}),
]
SEX_HEADER_CANDIDATES = {
    "sex",
    "gender",
    "child_sex",
    "child_gender",
    "sex_of_child",
    "gender_of_child",
}


def normalize_code(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header_key(value: object) -> str:
    text = normalize_code(value).casefold()
    if not text:
        return ""
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def normalize_date(value: object):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def normalize_age_months(value: object):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = normalize_code(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def quarter_label(value: date) -> str:
    quarter = ((value.month - 1) // 3) + 1
    return f"Q{quarter} {value.year}"


def calculate_completed_dose(age_months, source_values: dict[str, str], date_values: dict[str, date | None]) -> str:
    if age_months is None:
        return "not complete yet"

    if 0 <= age_months <= 11:
        group_label = "U1"
        required_doses = U1_REQUIRED_DOSES
    elif 12 <= age_months <= 59:
        group_label = "1-5"
        required_doses = ONE_TO_FIVE_REQUIRED_DOSES
    else:
        return "not complete yet"

    if any(source_values.get(dose, "Not received yet") == "Not received yet" for dose in required_doses):
        return "not complete yet"

    kdhw_dates = [
        date_values.get(dose)
        for dose in COMPLETION_CHECK_DOSES
        if source_values.get(dose) == "KDHW" and date_values.get(dose) is not None
    ]
    if not kdhw_dates:
        return "not complete yet"

    last_kdhw_date = max(kdhw_dates)
    return f"{group_label} completed in {quarter_label(last_kdhw_date)}"


def has_received(source_value: str | None) -> bool:
    return normalize_code(source_value) not in {"", "Not received yet"}


def find_unlogical_sequence_issues(source_values: dict[str, str], date_values: dict[str, date | None]) -> list[str]:
    issues: list[str] = []

    for group_name, doses in SEQUENCE_RULES.items():
        received_flags = [has_received(source_values.get(dose)) for dose in doses]

        for idx, dose in enumerate(doses[:-1]):
            if not received_flags[idx] and any(received_flags[idx + 1 :]):
                later_doses = ", ".join(doses[idx + 1 :])
                issues.append(f"{dose} not received but later dose(s) received ({later_doses})")

        for earlier_idx, earlier_dose in enumerate(doses[:-1]):
            earlier_date = date_values.get(earlier_dose)
            if earlier_date is None:
                continue

            for later_dose in doses[earlier_idx + 1 :]:
                later_date = date_values.get(later_dose)
                if later_date is not None and later_date < earlier_date:
                    issues.append(f"{later_dose} date earlier than {earlier_dose} date")

    return issues


def datediff_months(start_date: date | None, end_date: date | None):
    if start_date is None or end_date is None:
        return None
    if end_date < start_date:
        return None

    months = (end_date.year - start_date.year) * 12 + (end_date.month - start_date.month)
    if end_date.day < start_date.day:
        months -= 1
    return months


def build_source_value(date_value: object, other_value: object) -> str:
    if normalize_code(other_value).casefold() == "yes":
        return "Other"
    if normalize_date(date_value) is not None:
        return "KDHW"
    return "Not received yet"


def period_from_date(value: date) -> str:
    quarter = ((value.month - 1) // 3) + 1
    return f"Q{quarter}_{value.year}"


def parse_completed_period(value: str) -> tuple[str, str] | None:
    match = re.match(r"^(U1|1-5)\s+completed\s+in\s+Q([1-4])[ _](\d{4})$", normalize_code(value))
    if not match:
        return None
    group = match.group(1)
    period = f"Q{match.group(2)}_{match.group(3)}"
    return group, period


def age_bucket(age_months: int | None):
    if age_months is None:
        return None
    if 0 <= age_months <= 11:
        return "U1"
    if 12 <= age_months <= 59:
        return "U5"
    if age_months >= 60:
        return ">5"
    return None


def quarter_number(value: date) -> int:
    return ((value.month - 1) // 3) + 1


def sex_bucket(value: object) -> str | None:
    text = normalize_code(value).casefold()
    if not text:
        return None
    if text in {"m", "male", "boy", "boys"}:
        return "Male"
    if text in {"f", "female", "girl", "girls"}:
        return "Female"
    return None


def find_first_matching_header(headers: list[object], candidates: set[str]) -> str | None:
    for header in headers:
        if isinstance(header, str) and normalize_header_key(header) in candidates:
            return header
    return None


def dose_u1_eligible(dose_key: str, age_at_dose: int | None) -> bool:
    if age_at_dose is None:
        return False
    if not (0 <= age_at_dose <= 11):
        return False

    if dose_key in {"OPV1", "Penta1"}:
        return age_at_dose >= 2
    if dose_key in {"OPV2", "Penta2"}:
        return age_at_dose >= 3
    if dose_key in {"OPV3", "Penta3"}:
        return age_at_dose >= 4
    if dose_key == "MMR1":
        return age_at_dose >= 9
    if dose_key == "MMR2":
        return age_at_dose >= 10
    return True


def dose_group_key(vaccine_dose: str) -> str | None:
    if vaccine_dose in {"BCG", "OPV1", "OPV2", "OPV3", "Penta1", "Penta2", "Penta3", "MMR1", "MMR2", "IPV"}:
        return vaccine_dose
    if vaccine_dose in {"JE1", "JE2"}:
        return "JE"
    return None


def init_disaggregate_bucket() -> dict[str, set[str]]:
    return {column: set() for column in DISAGGREGATE_COUNT_COLUMNS}


def init_yearly_cumulative_bucket() -> dict[str, set[str]]:
    return {
        "ALOD-U1": set(),
        "ALOD-U5": set(),
        "ALOD->5": set(),
        "Td At least one dose": set(),
    }


def parse_year_value(value: object):
    text = normalize_code(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def build_project_lookup(source_workbook: Workbook) -> dict[tuple[int, str], str]:
    reference_sheet_name = None
    for name in source_workbook.sheetnames:
        if normalize_code(name).casefold() == "reference":
            reference_sheet_name = name
            break
    if reference_sheet_name is None:
        for name in source_workbook.sheetnames:
            if "reference" in normalize_code(name).casefold():
                reference_sheet_name = name
                break
    if reference_sheet_name is None:
        return {}

    ws = source_workbook[reference_sheet_name]
    headers = [normalize_code(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    normalized_headers = [normalize_header_key(header) for header in headers]

    year_candidates = {"year", "reporting_year", "yr"}
    clinic_candidates = {"clinic", "clinics", "clinic_name", "clinic_names", "vthc", "vthc_name"}
    project_candidates = {"project", "project_name"}

    year_col = next((i + 1 for i, h in enumerate(normalized_headers) if h in year_candidates), None)
    clinic_col = next((i + 1 for i, h in enumerate(normalized_headers) if h in clinic_candidates), None)
    project_col = next((i + 1 for i, h in enumerate(normalized_headers) if h in project_candidates), None)
    if year_col is None or clinic_col is None or project_col is None:
        return {}

    lookup: dict[tuple[int, str], str] = {}
    for r in range(2, ws.max_row + 1):
        year_value = parse_year_value(ws.cell(r, year_col).value)
        clinic_value = normalize_code(ws.cell(r, clinic_col).value)
        project_value = normalize_code(ws.cell(r, project_col).value)
        if year_value is None or not clinic_value or not project_value:
            continue
        lookup[(year_value, clinic_value.casefold())] = project_value

    return lookup


def build_projects_by_year(lookup: dict[tuple[int, str], str]) -> dict[int, list[str]]:
    projects_by_year: dict[int, set[str]] = {}
    for (year, _clinic), project_name in lookup.items():
        if not project_name:
            continue
        projects_by_year.setdefault(year, set()).add(project_name)
    return {year: sorted(project_names) for year, project_names in projects_by_year.items()}


def project_name_for_row(year: int, clinic: str, lookup: dict[tuple[int, str], str]) -> str:
    if year == 2023:
        return "RISE"
    if year == 2024:
        return "REACH-KK"
    if year in {2025, 2026}:
        return lookup.get((year, normalize_code(clinic).casefold()), "")
    return ""


def init_indicator_sex_bucket() -> dict[str, set[str]]:
    return {
        "U1 Male": set(),
        "U1 Female": set(),
        "1-5 Male": set(),
        "1-5 Female": set(),
    }


def init_indicator_counts() -> dict[str, dict[int, dict[str, set[str]]]]:
    return {
        label: {1: init_indicator_sex_bucket(), 2: init_indicator_sex_bucket(), 3: init_indicator_sex_bucket(), 4: init_indicator_sex_bucket()}
        for label, _metric, _ages in INDICATOR_DEFINITIONS
    }


def init_cummu_indicator_counts() -> dict[str, dict[str, set[str]]]:
    return {
        label: init_indicator_sex_bucket()
        for label, _metric, _ages in INDICATOR_DEFINITIONS
    }


def build_indicator_sheet(output: Workbook, source_workbook: Workbook) -> None:
    if LONG_FORM_SHEET not in output.sheetnames:
        return

    child_ws = output[LONG_FORM_SHEET]
    report = output.create_sheet(title=INDICATOR_SHEET)
    report.append(INDICATOR_HEADERS)

    child_headers = [child_ws.cell(1, c).value for c in range(1, child_ws.max_column + 1)]
    child_idx = {name: child_headers.index(name) + 1 for name in child_headers if isinstance(name, str)}
    sex_header = find_first_matching_header(child_headers, SEX_HEADER_CANDIDATES)
    if sex_header is None:
        raise KeyError("Could not find a sex/gender column for indicators. Expected one of: sex, gender, child_sex, child_gender, sex_of_child, gender_of_child.")
    sex_col = child_idx[sex_header]

    project_lookup = build_project_lookup(source_workbook)
    projects_by_year = build_projects_by_year(project_lookup)
    aggregate: dict[tuple[int, str, str], dict[str, dict[int, dict[str, set[str]]]]] = {}

    for r in range(2, child_ws.max_row + 1):
        children_code = normalize_code(child_ws.cell(r, child_idx["children_code"]).value)
        reporting_month = normalize_date(child_ws.cell(r, child_idx["reporting_month"]).value)
        source_value = normalize_code(child_ws.cell(r, child_idx["source"]).value)
        vaccine_dose = normalize_code(child_ws.cell(r, child_idx["vaccine_dose"]).value)
        if not children_code or reporting_month is None or source_value != "KDHW":
            continue

        sex_value = sex_bucket(child_ws.cell(r, sex_col).value) if sex_col is not None else None
        if sex_value is None:
            continue

        year = reporting_month.year
        quarter = quarter_number(reporting_month)
        clinic = normalize_code(child_ws.cell(r, child_idx["vthc_name"]).value)
        project_name = project_name_for_row(year, clinic, project_lookup)
        aggregate_key = (year, "KDHW", project_name)
        indicator_counts = aggregate.setdefault(aggregate_key, init_indicator_counts())

        age_at_dose = normalize_age_months(child_ws.cell(r, child_idx["age_at_dose"]).value)
        dose_age = age_bucket(age_at_dose)
        if dose_age in {"U1", "U5"}:
            for label, metric, age_groups in INDICATOR_DEFINITIONS:
                if metric != vaccine_dose or dose_age not in age_groups:
                    continue
                age_label = "U1" if dose_age == "U1" else "1-5"
                indicator_counts[label][quarter][f"{age_label} {sex_value}"].add(children_code)

        age_months = normalize_age_months(child_ws.cell(r, child_idx["ageInMonths"]).value)
        current_age_group = age_bucket(age_months)
        if current_age_group in {"U1", "U5"}:
            current_age_label = "U1" if current_age_group == "U1" else "1-5"
            indicator_counts["At least one dose under 5-yr-old"][quarter][f"{current_age_label} {sex_value}"].add(children_code)

        completed_value = normalize_code(child_ws.cell(r, child_idx["completed_dose"]).value)
        completed_info = parse_completed_period(completed_value)
        if completed_info:
            completed_group, completed_period = completed_info
            completed_quarter_match = re.match(r"Q([1-4])_(\d{4})", completed_period)
            if completed_quarter_match:
                completed_quarter = int(completed_quarter_match.group(1))
                completed_year = int(completed_quarter_match.group(2))
                completed_project_name = project_name_for_row(completed_year, clinic, project_lookup)
                completed_key = (completed_year, "KDHW", completed_project_name)
                completed_counts = aggregate.setdefault(completed_key, init_indicator_counts())
                completed_age_label = "U1" if completed_group == "U1" else "1-5"
                completed_counts["Full dose under 5-yr-old"][completed_quarter][f"{completed_age_label} {sex_value}"].add(children_code)

    output_keys = set(aggregate.keys())
    for year in (2025, 2026):
        for project_name in projects_by_year.get(year, []):
            output_keys.add((year, "KDHW", project_name))

    for year, organization, project_name in sorted(output_keys):
        indicator_counts = aggregate.setdefault((year, organization, project_name), init_indicator_counts())
        for label, _metric, _age_groups in INDICATOR_DEFINITIONS:
            row = [year, organization, project_name, label]
            for quarter in range(1, 5):
                quarter_counts = indicator_counts[label][quarter]
                u1_male = len(quarter_counts["U1 Male"])
                u1_female = len(quarter_counts["U1 Female"])
                one_to_five_male = len(quarter_counts["1-5 Male"])
                one_to_five_female = len(quarter_counts["1-5 Female"])
                total = u1_male + u1_female + one_to_five_male + one_to_five_female
                row.extend(["", u1_male, u1_female, one_to_five_male, one_to_five_female, total])
            report.append(row)


def build_cummulative_sheet(output: Workbook, source_workbook: Workbook) -> None:
    if LONG_FORM_SHEET not in output.sheetnames or PW_LONG_FORM_SHEET not in output.sheetnames:
        return

    child_ws = output[LONG_FORM_SHEET]
    pw_ws = output[PW_LONG_FORM_SHEET]
    report = output.create_sheet(title=CUMULATIVE_SHEET)
    report.append(YEARLY_CUMULATIVE_HEADERS)

    child_headers = [child_ws.cell(1, c).value for c in range(1, child_ws.max_column + 1)]
    child_idx = {name: child_headers.index(name) + 1 for name in child_headers if isinstance(name, str)}

    pw_headers = [pw_ws.cell(1, c).value for c in range(1, pw_ws.max_column + 1)]
    pw_idx = {name: pw_headers.index(name) + 1 for name in pw_headers if isinstance(name, str)}

    project_lookup = build_project_lookup(source_workbook)
    aggregate: dict[tuple[str, int, str, str, str, str, str], dict[str, set[str]]] = {}

    for r in range(2, child_ws.max_row + 1):
        children_code = normalize_code(child_ws.cell(r, child_idx["children_code"]).value)
        reporting_month = normalize_date(child_ws.cell(r, child_idx["reporting_month"]).value)
        source_value = normalize_code(child_ws.cell(r, child_idx["source"]).value)
        if not children_code or reporting_month is None or source_value != "KDHW":
            continue

        year = reporting_month.year
        district = normalize_code(child_ws.cell(r, child_idx["district"]).value)
        township_eho = normalize_code(child_ws.cell(r, child_idx["township_name"]).value)
        twp_mimu = normalize_code(child_ws.cell(r, child_idx["township_name_MIMU"]).value)
        clinic = normalize_code(child_ws.cell(r, child_idx["vthc_name"]).value)
        project_name = project_name_for_row(year, clinic, project_lookup)
        key = ("KDHW", year, project_name, district, township_eho, twp_mimu, clinic)
        bucket = aggregate.setdefault(key, init_yearly_cumulative_bucket())

        age_months = normalize_age_months(child_ws.cell(r, child_idx["ageInMonths"]).value)
        alod = age_bucket(age_months)
        if alod == "U1":
            bucket["ALOD-U1"].add(children_code)
        elif alod == "U5":
            bucket["ALOD-U5"].add(children_code)
        elif alod == ">5":
            bucket["ALOD->5"].add(children_code)

    for r in range(2, pw_ws.max_row + 1):
        mother_code = normalize_code(pw_ws.cell(r, pw_idx["mother_code"]).value)
        reporting_month = normalize_date(pw_ws.cell(r, pw_idx["reporting_month"]).value)
        vaccine_dose = normalize_code(pw_ws.cell(r, pw_idx["vaccine_dose"]).value)
        source_value = normalize_code(pw_ws.cell(r, pw_idx["source"]).value)
        if not mother_code or reporting_month is None or source_value != "KDHW":
            continue

        year = reporting_month.year
        district = normalize_code(pw_ws.cell(r, pw_idx["district"]).value)
        township_eho = normalize_code(pw_ws.cell(r, pw_idx["township_name"]).value)
        twp_mimu = normalize_code(pw_ws.cell(r, pw_idx["township_name_MIMU"]).value)
        clinic = normalize_code(pw_ws.cell(r, pw_idx["vthc_name"]).value)
        project_name = project_name_for_row(year, clinic, project_lookup)
        key = ("KDHW", year, project_name, district, township_eho, twp_mimu, clinic)
        bucket = aggregate.setdefault(key, init_yearly_cumulative_bucket())
        if vaccine_dose in {"Td1", "Td2"}:
            bucket["Td At least one dose"].add(mother_code)

    def sort_key(value: tuple[str, int, str, str, str, str, str]):
        _org, year, _project, district, township, twp_mimu, clinic = value
        return (year, district, township, twp_mimu, clinic)

    for key in sorted(aggregate.keys(), key=sort_key):
        org, year, project, district, township, twp_mimu, clinic = key
        bucket = aggregate[key]
        report.append(
            [
                org,
                year,
                project,
                district,
                township,
                twp_mimu,
                clinic,
                len(bucket["ALOD-U1"]),
                len(bucket["ALOD-U5"]),
                len(bucket["ALOD->5"]),
                len(bucket["Td At least one dose"]),
            ]
        )


def build_cummu_indicator_sheet(output: Workbook, source_workbook: Workbook) -> None:
    if LONG_FORM_SHEET not in output.sheetnames:
        return

    child_ws = output[LONG_FORM_SHEET]
    report = output.create_sheet(title=CUMMU_INDICATOR_SHEET)
    report.append(CUMMU_INDICATOR_HEADERS)

    child_headers = [child_ws.cell(1, c).value for c in range(1, child_ws.max_column + 1)]
    child_idx = {name: child_headers.index(name) + 1 for name in child_headers if isinstance(name, str)}
    sex_header = find_first_matching_header(child_headers, SEX_HEADER_CANDIDATES)
    if sex_header is None:
        raise KeyError("Could not find a sex/gender column for indicators. Expected one of: sex, gender, child_sex, child_gender, sex_of_child, gender_of_child.")
    sex_col = child_idx[sex_header]

    project_lookup = build_project_lookup(source_workbook)
    projects_by_year = build_projects_by_year(project_lookup)
    aggregate: dict[tuple[int, str, str], dict[str, dict[str, set[str]]]] = {}

    for r in range(2, child_ws.max_row + 1):
        children_code = normalize_code(child_ws.cell(r, child_idx["children_code"]).value)
        reporting_month = normalize_date(child_ws.cell(r, child_idx["reporting_month"]).value)
        source_value = normalize_code(child_ws.cell(r, child_idx["source"]).value)
        vaccine_dose = normalize_code(child_ws.cell(r, child_idx["vaccine_dose"]).value)
        if not children_code or reporting_month is None or source_value != "KDHW":
            continue

        sex_value = sex_bucket(child_ws.cell(r, sex_col).value)
        if sex_value is None:
            continue

        year = reporting_month.year
        clinic = normalize_code(child_ws.cell(r, child_idx["vthc_name"]).value)
        project_name = project_name_for_row(year, clinic, project_lookup)
        aggregate_key = (year, "KDHW", project_name)
        indicator_counts = aggregate.setdefault(aggregate_key, init_cummu_indicator_counts())

        age_at_dose = normalize_age_months(child_ws.cell(r, child_idx["age_at_dose"]).value)
        dose_age = age_bucket(age_at_dose)
        if dose_age in {"U1", "U5"}:
            for label, metric, age_groups in INDICATOR_DEFINITIONS:
                if metric != vaccine_dose or dose_age not in age_groups:
                    continue
                age_label = "U1" if dose_age == "U1" else "1-5"
                indicator_counts[label][f"{age_label} {sex_value}"].add(children_code)

        age_months = normalize_age_months(child_ws.cell(r, child_idx["ageInMonths"]).value)
        current_age_group = age_bucket(age_months)
        if current_age_group in {"U1", "U5"}:
            current_age_label = "U1" if current_age_group == "U1" else "1-5"
            indicator_counts["At least one dose under 5-yr-old"][f"{current_age_label} {sex_value}"].add(children_code)

        completed_value = normalize_code(child_ws.cell(r, child_idx["completed_dose"]).value)
        completed_info = parse_completed_period(completed_value)
        if completed_info:
            completed_group, completed_period = completed_info
            completed_year = int(completed_period.split("_")[1])
            completed_project_name = project_name_for_row(completed_year, clinic, project_lookup)
            completed_key = (completed_year, "KDHW", completed_project_name)
            completed_counts = aggregate.setdefault(completed_key, init_cummu_indicator_counts())
            completed_age_label = "U1" if completed_group == "U1" else "1-5"
            completed_counts["Full dose under 5-yr-old"][f"{completed_age_label} {sex_value}"].add(children_code)

    output_keys = set(aggregate.keys())
    for year in (2025, 2026):
        for project_name in projects_by_year.get(year, []):
            output_keys.add((year, "KDHW", project_name))

    for year, organization, project_name in sorted(output_keys):
        indicator_counts = aggregate.setdefault((year, organization, project_name), init_cummu_indicator_counts())
        for label, _metric, _age_groups in INDICATOR_DEFINITIONS:
            counts = indicator_counts[label]
            report.append(
                [
                    year,
                    organization,
                    project_name,
                    label,
                    "",
                    len(counts["U1 Male"]),
                    len(counts["U1 Female"]),
                    len(counts["1-5 Male"]),
                    len(counts["1-5 Female"]),
                ]
            )


def build_vthc_dose_disaggregate_sheet(output: Workbook, source_workbook: Workbook) -> None:
    if LONG_FORM_SHEET not in output.sheetnames or PW_LONG_FORM_SHEET not in output.sheetnames:
        return

    child_ws = output[LONG_FORM_SHEET]
    pw_ws = output[PW_LONG_FORM_SHEET]
    report = output.create_sheet(title=VTHC_DISAGGREGATE_SHEET)
    report.append(DISAGGREGATE_HEADERS)

    child_headers = [child_ws.cell(1, c).value for c in range(1, child_ws.max_column + 1)]
    child_idx = {name: child_headers.index(name) + 1 for name in child_headers if isinstance(name, str)}

    pw_headers = [pw_ws.cell(1, c).value for c in range(1, pw_ws.max_column + 1)]
    pw_idx = {name: pw_headers.index(name) + 1 for name in pw_headers if isinstance(name, str)}

    project_lookup = build_project_lookup(source_workbook)
    aggregate: dict[tuple[int, str, str, str, str, str, str, str], dict[str, set[str]]] = {}

    for r in range(2, child_ws.max_row + 1):
        children_code = normalize_code(child_ws.cell(r, child_idx["children_code"]).value)
        reporting_month = normalize_date(child_ws.cell(r, child_idx["reporting_month"]).value)
        vaccine_dose = normalize_code(child_ws.cell(r, child_idx["vaccine_dose"]).value)
        source_value = normalize_code(child_ws.cell(r, child_idx["source"]).value)
        if not children_code or reporting_month is None or source_value != "KDHW":
            continue

        period = period_from_date(reporting_month)
        year = reporting_month.year
        district = normalize_code(child_ws.cell(r, child_idx["district"]).value)
        township_eho = normalize_code(child_ws.cell(r, child_idx["township_name"]).value)
        twp_mimu = normalize_code(child_ws.cell(r, child_idx["township_name_MIMU"]).value)
        clinic = normalize_code(child_ws.cell(r, child_idx["vthc_name"]).value)

        project_name = project_name_for_row(year, clinic, project_lookup)
        key = (year, period, "KDHW", project_name, district, township_eho, twp_mimu, clinic)
        bucket = aggregate.setdefault(key, init_disaggregate_bucket())

        age_months = normalize_age_months(child_ws.cell(r, child_idx["ageInMonths"]).value)
        alod = age_bucket(age_months)
        if alod == "U1":
            bucket["ALOD-U1"].add(children_code)
        elif alod == "U5":
            bucket["ALOD-U5"].add(children_code)
        elif alod == ">5":
            bucket["ALOD->5"].add(children_code)

        dose_group = dose_group_key(vaccine_dose)
        age_at_dose = normalize_age_months(child_ws.cell(r, child_idx["age_at_dose"]).value)
        dose_age = age_bucket(age_at_dose)
        if dose_group and dose_age:
            if dose_age == "U1":
                if dose_u1_eligible(dose_group, age_at_dose):
                    bucket[f"{dose_group}_U1"].add(children_code)
            elif dose_age == "U5":
                bucket[f"{dose_group}_U5"].add(children_code)
            elif dose_age == ">5":
                bucket[f"{dose_group}_>5"].add(children_code)

        completed_value = normalize_code(child_ws.cell(r, child_idx["completed_dose"]).value)
        completed_info = parse_completed_period(completed_value)
        if completed_info:
            completed_group, completed_period = completed_info
            completed_year = int(completed_period.split("_")[1])
            completed_project_name = project_name_for_row(completed_year, clinic, project_lookup)
            completed_key = (completed_year, completed_period, "KDHW", completed_project_name, district, township_eho, twp_mimu, clinic)
            completed_bucket = aggregate.setdefault(completed_key, init_disaggregate_bucket())
            if completed_group == "U1":
                completed_bucket["CD_U1"].add(children_code)
            elif completed_group == "1-5":
                completed_bucket["CD_U5"].add(children_code)

    for r in range(2, pw_ws.max_row + 1):
        mother_code = normalize_code(pw_ws.cell(r, pw_idx["mother_code"]).value)
        reporting_month = normalize_date(pw_ws.cell(r, pw_idx["reporting_month"]).value)
        vaccine_dose = normalize_code(pw_ws.cell(r, pw_idx["vaccine_dose"]).value)
        source_value = normalize_code(pw_ws.cell(r, pw_idx["source"]).value)
        if not mother_code or reporting_month is None or source_value != "KDHW":
            continue

        period = period_from_date(reporting_month)
        year = reporting_month.year
        district = normalize_code(pw_ws.cell(r, pw_idx["district"]).value)
        township_eho = normalize_code(pw_ws.cell(r, pw_idx["township_name"]).value)
        twp_mimu = normalize_code(pw_ws.cell(r, pw_idx["township_name_MIMU"]).value)
        clinic = normalize_code(pw_ws.cell(r, pw_idx["vthc_name"]).value)

        project_name = project_name_for_row(year, clinic, project_lookup)
        key = (year, period, "KDHW", project_name, district, township_eho, twp_mimu, clinic)
        bucket = aggregate.setdefault(key, init_disaggregate_bucket())
        if vaccine_dose == "Td1":
            bucket["Td1"].add(mother_code)
        if vaccine_dose == "Td2":
            bucket["Td2"].add(mother_code)
        if vaccine_dose in {"Td1", "Td2"}:
            bucket["Td At least one dose"].add(mother_code)

    def sort_key(value: tuple[int, str, str, str, str, str, str, str]):
        year, period, _org, _project, district, township, twp_mimu, clinic = value
        q_match = re.match(r"Q([1-4])_(\d{4})", period)
        quarter = int(q_match.group(1)) if q_match else 0
        return (year, quarter, district, township, twp_mimu, clinic)

    for key in sorted(aggregate.keys(), key=sort_key):
        year, period, org, project, district, township, twp_mimu, clinic = key
        bucket = aggregate[key]
        row = [
            year,
            period,
            org,
            project,
            district,
            township,
            twp_mimu,
            clinic,
        ]
        row.extend(len(bucket[column]) for column in DISAGGREGATE_COUNT_COLUMNS)
        report.append(row)


def should_exclude_output_column(header: object) -> bool:
    return isinstance(header, str) and header in EXCLUDED_OUTPUT_HEADERS


def build_output_columns(headers: list[object]) -> tuple[list[int], list[str], dict[int, int]]:
    included_columns: list[int] = []
    output_headers: list[str] = []
    source_column_map: dict[int, int] = {}

    other_to_source = {other_header: source_header for _, other_header, source_header in SOURCE_FIELD_RULES}

    for source_index, header in enumerate(headers, start=1):
        if should_exclude_output_column(header):
            continue

        if isinstance(header, str) and header in other_to_source:
            included_columns.append(source_index)
            source_column_map[source_index] = len(output_headers) + 1
            output_headers.append(other_to_source[header])
            continue

        included_columns.append(source_index)
        source_column_map[source_index] = len(output_headers) + 1
        output_headers.append(header if isinstance(header, str) else "")

    return included_columns, output_headers, source_column_map


def load_source_workbook(uploaded_file):
    if uploaded_file is None:
        raise ValueError("Please upload a workbook to verify.")
    return load_workbook(uploaded_file), uploaded_file.name


def get_children_sheet(workbook: Workbook):
    if SOURCE_SHEET not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise KeyError(f"Sheet '{SOURCE_SHEET}' was not found. Available sheets: {available}")
    return workbook[SOURCE_SHEET]


def get_pw_sheet(workbook: Workbook):
    if PW_SHEET not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise KeyError(f"Sheet '{PW_SHEET}' was not found. Available sheets: {available}")
    return workbook[PW_SHEET]


def build_pw_verification_sheet(output: Workbook, pw_sheet) -> dict[str, int]:
    headers = [pw_sheet.cell(1, column).value for column in range(1, pw_sheet.max_column + 1)]
    if PW_CODE_HEADER not in headers:
        raise KeyError(f"Column '{PW_CODE_HEADER}' was not found in sheet '{PW_SHEET}'.")

    mother_code_column = headers.index(PW_CODE_HEADER) + 1
    rows = list(pw_sheet.iter_rows(min_row=2, values_only=True))

    normalized_codes = [normalize_code(row[mother_code_column - 1]) for row in rows]
    non_empty_codes = [code for code in normalized_codes if code]
    code_counts = Counter(non_empty_codes)
    duplicate_codes = {code for code, count in code_counts.items() if count > 1}

    report = output.create_sheet(title=PW_VERIFICATION_SHEET)
    pw_long = output.create_sheet(title=PW_LONG_FORM_SHEET)

    present_source_rules = [rule for rule in PW_SOURCE_FIELD_RULES if rule[0] in headers and rule[1] in headers]
    other_to_source = {other_header: source_header for _, other_header, source_header in present_source_rules}

    included_columns: list[int] = []
    output_headers: list[str] = []
    report_column_map: dict[int, int] = {}

    for source_index, header in enumerate(headers, start=1):
        if isinstance(header, str) and header in PW_DROP_HEADERS:
            continue

        included_columns.append(source_index)
        report_column_map[source_index] = len(output_headers) + 1
        if isinstance(header, str) and header in other_to_source:
            output_headers.append(other_to_source[header])
        else:
            output_headers.append(header if isinstance(header, str) else "")

    report.append(output_headers + ["verification_status"])

    if "IDP" not in headers:
        raise KeyError("Column 'IDP' was not found in sheet 'PW'.")
    idp_column = headers.index("IDP") + 1
    static_columns = list(range(1, idp_column + 1))
    pw_long_headers = [headers[index - 1] for index in static_columns] + [
        "vaccine_dose",
        "reporting_month",
        "source",
    ]
    pw_long.append(pw_long_headers)

    missing_row_count = 0
    duplicate_row_count = 0
    unlogical_row_count = 0
    affected_row_count = 0

    rule_by_other_header = {rule[1]: rule for rule in present_source_rules}

    for row_index, row in enumerate(rows, start=2):
        row_values = list(row)
        mother_code = normalize_code(row_values[mother_code_column - 1])
        status = "OK"
        issues: list[str] = []

        if not mother_code:
            status = "Missing mother_code"
            missing_row_count += 1
        elif mother_code in duplicate_codes:
            status = "Duplicate mother_code"
            duplicate_row_count += 1

        output_row_values: list[object] = []
        td_source_values: dict[str, str] = {}

        for source_index in included_columns:
            header = headers[source_index - 1]
            if isinstance(header, str) and header in rule_by_other_header:
                date_header, other_header, source_header = rule_by_other_header[header]
                date_value = row_values[headers.index(date_header)]
                other_value = row_values[headers.index(other_header)]
                source_value = build_source_value(date_value, other_value)
                output_row_values.append(source_value)
                if source_header == "Td1_source":
                    td_source_values["TD1"] = source_value
                elif source_header == "Td2_source":
                    td_source_values["TD2"] = source_value
            else:
                output_row_values.append(row_values[source_index - 1])

        td1_date = normalize_date(row_values[headers.index("td_first_dose")]) if "td_first_dose" in headers else None
        td2_date = normalize_date(row_values[headers.index("td_second_dose")]) if "td_second_dose" in headers else None
        td1_received = has_received(td_source_values.get("TD1"))
        td2_received = has_received(td_source_values.get("TD2"))

        if not td1_received and td2_received:
            issues.append("Td1 not received but Td2 received")
        if td1_date is not None and td2_date is not None and td2_date < td1_date:
            issues.append("Td2 date earlier than Td1 date")

        if issues:
            unlogical_row_count += 1
            if status == "OK":
                status = "; ".join(issues)
            else:
                status = f"{status}; " + "; ".join(issues)

        report.append(output_row_values + [status])

        static_values = [row_values[index - 1] for index in static_columns]
        pw_source_by_dose = {
            "Td1": td_source_values.get("TD1", "Not received yet"),
            "Td2": td_source_values.get("TD2", "Not received yet"),
        }
        for dose_key in PW_LONG_DOSES:
            source_value = pw_source_by_dose.get(dose_key, "Not received yet")
            if source_value == "Not received yet":
                continue

            reporting_header = PW_DOSE_REPORTING_HEADERS.get(dose_key)
            reporting_month_value = row_values[headers.index(reporting_header)] if reporting_header in headers else None
            pw_long.append(static_values + [dose_key, reporting_month_value, source_value])

        if status != "OK":
            affected_row_count += 1
            code_cell = report.cell(row=row_index, column=report_column_map[mother_code_column])
            code_cell.fill = RED_FILL
            code_cell.font = RED_FONT

    summary = {
        "rows": len(rows),
        "missing_rows": missing_row_count,
        "duplicate_rows": duplicate_row_count,
        "duplicate_codes": len(duplicate_codes),
        "unlogical_rows": unlogical_row_count,
        "affected_rows": affected_row_count,
    }
    return summary


def build_verification_report(source_sheet) -> tuple[Workbook, dict[str, int]]:
    headers = [source_sheet.cell(1, column).value for column in range(1, source_sheet.max_column + 1)]
    if CODE_HEADER not in headers:
        raise KeyError(f"Column '{CODE_HEADER}' was not found in sheet '{SOURCE_SHEET}'.")
    if "date_of_birth" not in headers:
        raise KeyError("Column 'date_of_birth' was not found in sheet 'children'.")

    code_column = headers.index(CODE_HEADER) + 1
    dob_column = headers.index("date_of_birth") + 1
    registered_column = headers.index("registered_date") + 1 if "registered_date" in headers else None
    comparison_columns = [headers.index(name) + 1 for name in DATE_COLUMNS if name in headers and name != "registered_date"]
    rows = list(source_sheet.iter_rows(min_row=2, values_only=True))
    normalized_codes = [normalize_code(row[code_column - 1]) for row in rows]
    non_empty_codes = [code for code in normalized_codes if code]
    code_counts = Counter(non_empty_codes)
    duplicate_codes = {code for code, count in code_counts.items() if count > 1}

    output = Workbook()
    report = output.active
    report.title = VERIFICATION_SHEET
    long_form = output.create_sheet(title=LONG_FORM_SHEET)

    included_columns, report_headers, report_column_map = build_output_columns(headers)
    present_source_rules = [rule for rule in SOURCE_FIELD_RULES if rule[0] in headers and rule[1] in headers]
    source_header_to_dose = {source_header: dose_key for dose_key, source_header in DOSE_SOURCE_HEADERS.items()}
    other_header_to_rule = {rule[1]: rule for rule in present_source_rules}

    report_headers = report_headers + ["completed_dose", "verification_status"]
    report.append(report_headers)

    if "IDP" not in headers:
        raise KeyError("Column 'IDP' was not found in sheet 'children'.")
    idp_column = headers.index("IDP") + 1
    static_columns = list(range(1, idp_column + 1))
    long_form_headers = [headers[index - 1] for index in static_columns] + [
        "completed_dose",
        "vaccine_dose",
        "reporting_month",
        "source",
        "age_at_dose",
    ]
    long_form.append(long_form_headers)

    missing_count = 0
    duplicate_row_count = 0
    dob_order_row_count = 0
    unlogical_row_count = 0
    affected_row_count = 0

    for row_index, row in enumerate(rows, start=2):
        row_values = list(row)
        code_value = normalize_code(row_values[code_column - 1])
        dob_value = normalize_date(row_values[dob_column - 1])
        status = "OK"
        dob_issues: list[str] = []

        if not code_value:
            status = "Missing children_code"
            missing_count += 1
        elif code_value in duplicate_codes:
            status = "Duplicate children_code"
            duplicate_row_count += 1

        if dob_value is not None and registered_column is not None:
            registered_value = normalize_date(row_values[registered_column - 1])
            if registered_value is not None and dob_value > registered_value:
                dob_issues.append("date_of_birth later than registered_date")

        if dob_value is not None:
            for column in comparison_columns:
                compared_value = normalize_date(row_values[column - 1])
                if compared_value is not None and dob_value > compared_value:
                    dob_issues.append(f"date_of_birth later than {headers[column - 1]}")

        output_row_values: list[object] = []
        source_values_by_dose: dict[str, str] = {}
        for index in included_columns:
            header = headers[index - 1]
            if isinstance(header, str) and header in other_header_to_rule:
                date_header, other_header, source_header = other_header_to_rule[header]
                date_value = row_values[headers.index(date_header)]
                other_value = row_values[headers.index(other_header)]
                source_value = build_source_value(date_value, other_value)
                output_row_values.append(source_value)
                dose_key = source_header_to_dose.get(source_header)
                if dose_key is not None:
                    source_values_by_dose[dose_key] = source_value
            else:
                output_row_values.append(row_values[index - 1])

        dose_date_values = {
            dose_key: normalize_date(row_values[headers.index(date_header)])
            for dose_key, date_header in DOSE_DATE_HEADERS.items()
            if date_header in headers
        }
        unlogical_issues = find_unlogical_sequence_issues(source_values_by_dose, dose_date_values)

        if dob_issues:
            dob_order_row_count += 1
        if unlogical_issues:
            unlogical_row_count += 1

        combined_issues = dob_issues + unlogical_issues
        if combined_issues:
            if status == "OK":
                status = "; ".join(combined_issues)
            else:
                status = f"{status}; " + "; ".join(combined_issues)

        age_months = normalize_age_months(row_values[headers.index("ageInMonths")]) if "ageInMonths" in headers else None
        completed_dose_value = calculate_completed_dose(age_months, source_values_by_dose, dose_date_values)

        report.append(output_row_values + [completed_dose_value, status])
        if status != "OK":
            affected_row_count += 1

        static_values = [row_values[index - 1] for index in static_columns]
        dob_for_age = normalize_date(row_values[dob_column - 1])
        for dose_key in LONG_FORM_DOSES:
            date_header = DOSE_DATE_HEADERS.get(dose_key)
            reporting_header = DOSE_REPORTING_HEADERS.get(dose_key)
            source_value = source_values_by_dose.get(dose_key, "Not received yet")
            if source_value == "Not received yet":
                continue

            dose_date_value = dose_date_values.get(dose_key)
            reporting_month_value = row_values[headers.index(reporting_header)] if reporting_header in headers else None
            age_at_dose = datediff_months(dob_for_age, dose_date_value)

            long_form.append(
                static_values
                + [
                    completed_dose_value,
                    dose_key,
                    reporting_month_value,
                    source_value,
                    age_at_dose,
                ]
            )

        if status != "OK":
            code_cell = report.cell(row=row_index, column=report_column_map[code_column])
            code_cell.fill = RED_FILL
            code_cell.font = RED_FONT

            dob_cell = report.cell(row=row_index, column=report_column_map[dob_column])
            if dob_issues:
                dob_cell.fill = RED_FILL
                dob_cell.font = RED_FONT

    duplicate_code_count = len(duplicate_codes)
    summary = {
        "rows": len(rows),
        "missing_rows": missing_count,
        "duplicate_rows": duplicate_row_count,
        "duplicate_codes": duplicate_code_count,
        "dob_order_rows": dob_order_row_count,
        "unlogical_rows": unlogical_row_count,
        "affected_rows": affected_row_count,
    }
    return output, summary


def workbook_to_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    st.title(APP_TITLE)
    st.write(
        "Upload a KDHW workbook or use the bundled sample file. The app checks the children sheet, "
        "flags missing or duplicate children_code values, and generates a verification workbook with red highlights."
    )

    uploaded_file = st.file_uploader("Upload an .xlsx workbook", type=["xlsx"])

    if uploaded_file is None:
        st.info("Upload a workbook to generate the verification file.")
        return

    try:
        source_workbook, source_path = load_source_workbook(uploaded_file)
        source_sheet = get_children_sheet(source_workbook)
        pw_sheet = get_pw_sheet(source_workbook)
        report_workbook, summary = build_verification_report(source_sheet)
        pw_summary = build_pw_verification_sheet(report_workbook, pw_sheet)
        build_vthc_dose_disaggregate_sheet(report_workbook, source_workbook)
        build_indicator_sheet(report_workbook, source_workbook)
        build_cummulative_sheet(report_workbook, source_workbook)
        build_cummu_indicator_sheet(report_workbook, source_workbook)
    except Exception as exc:
        st.error(str(exc))
        return

    st.caption(f"Using uploaded workbook: {source_path}")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Rows checked", summary["rows"])
    col2.metric("Missing children_code", summary["missing_rows"])
    col3.metric("Duplicate rows", summary["duplicate_rows"])
    col4.metric("Duplicate codes", summary["duplicate_codes"])

    st.metric("Rows with date order issues", summary["dob_order_rows"])
    st.metric("Rows with unlogical dose records", summary["unlogical_rows"])

    st.info(f"Rows with issues: {summary['affected_rows']}")
    if summary["unlogical_rows"] > 0:
        st.warning("Unlogical dose records detected for OPV, Penta, or MMR sequences. Check highlighted children_code rows.")

    st.subheader("PW sheet checks")
    pw_col1, pw_col2, pw_col3, pw_col4 = st.columns(4)
    pw_col1.metric("PW rows checked", pw_summary["rows"])
    pw_col2.metric("Missing mother_code rows", pw_summary["missing_rows"])
    pw_col3.metric("Duplicate mother_code rows", pw_summary["duplicate_rows"])
    pw_col4.metric("Duplicate mother_code values", pw_summary["duplicate_codes"])
    st.metric("PW rows with unlogical Td records", pw_summary["unlogical_rows"])
    st.info(f"PW rows with issues: {pw_summary['affected_rows']}")

    report_bytes = workbook_to_bytes(report_workbook)
    st.download_button(
        label="Download new verification file",
        data=report_bytes,
        file_name="KDHW_verification.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.subheader("Rules applied")
    st.write("Missing children_code values are flagged red.")
    st.write("Duplicate children_code values are flagged red on every matching row.")
    st.write("date_of_birth is checked against registered_date and the listed vaccine date columns; later values are flagged red.")
    st.write("Unlogical sequences are checked for OPV, Penta, and MMR: earlier dose missing while later dose received, or later dose date earlier than primary dose date.")
    st.write("PW sheet drops Td third/fourth/fifth groups, prevent_td_newborn, and comments; checks duplicate mother_code; replaces td_first_dose_other and td_second_dose_other with Td1_source and Td2_source.")
    st.write("PW unlogical checks: Td1 not received while Td2 received, and Td2 date earlier than Td1 date. Rows are flagged and mother_code is highlighted red.")
    st.write("Long-form sheets include only received doses; rows with source 'Not received yet' are excluded in children_long_form and pw_long.")
    st.write("VTHC_dose_disaggregate is built from children_long_form and pw_long with quarter period aggregation by clinic.")
    st.write("indicators is built from children_long_form with quarterly sex and age disaggregation for Penta1, Penta3, MMR1, MMR2, full dose, and at least one dose under 5.")
    st.write("cummulative_sheet is built from children_long_form and pw_long with yearly aggregation by clinic for ALOD and Td At least one dose.")
    st.write("Cummu_indicator is built from children_long_form with yearly sex and age disaggregation for the same indicator set as indicators.")


if __name__ == "__main__":
    main()
